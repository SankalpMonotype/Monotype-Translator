#!/usr/bin/env python3
"""
Translation comparison script.
Compares Linguist A vs AI B translations across 5 languages.
Output: ~/Downloads/Translation_Comparison.xlsx
"""

import os
import sys

# ── Install dependencies if missing ──────────────────────────────────────────
def ensure_package(pkg):
    try:
        __import__(pkg)
    except ImportError:
        import subprocess
        print(f"Installing {pkg}...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "--quiet"])

for pkg in ["rapidfuzz", "openpyxl", "pandas"]:
    ensure_package(pkg)

import pandas as pd
from rapidfuzz import fuzz
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────────────────
HOME = os.path.expanduser("~")
FILE_A = os.path.join(HOME, "Downloads", "Linguist A.xlsx")
FILE_B = os.path.join(HOME, "Downloads", "AI B.xlsx")
OUTPUT  = os.path.join(HOME, "Downloads", "Translation_Comparison.xlsx")

# Languages to compare (canonical names used as sheet names)
LANGUAGES = ["German", "French", "Spanish", "Portuguese", "Japanese"]

# Column-name maps per file  (lowercase for matching)
COL_MAP_A = {
    "german":     "German",
    "french":     "French",
    "spanish":    "Spanish",
    "portuguese": "Portuguese",
    "japanese":   "Japanese",
    "english":    "English",
}
COL_MAP_B = {
    "german":     "German",
    "french":     "French",
    "spanish":    "Spanish",
    "portuguese": "Portuguese",
    "japanese":   "Japanese",
    "english":    "English",
}

# ── Colour helpers ────────────────────────────────────────────────────────────
GREEN  = PatternFill("solid", fgColor="92D050")   # ≥ 80
YELLOW = PatternFill("solid", fgColor="FFEB84")   # 50–79
RED    = PatternFill("solid", fgColor="FF6B6B")   # < 50
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_FILL    = PatternFill("solid", fgColor="EEF3F8")

thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def score_fill(score: float) -> PatternFill:
    if score >= 80:
        return GREEN
    elif score >= 50:
        return YELLOW
    else:
        return RED


def auto_width(ws, col_idx: int, min_w: int = 10, max_w: int = 60):
    col_letter = get_column_letter(col_idx)
    best = min_w
    for cell in ws[col_letter]:
        if cell.value:
            best = max(best, min(len(str(cell.value)) + 2, max_w))
    ws.column_dimensions[col_letter].width = best


# ── Load data ─────────────────────────────────────────────────────────────────
print("Loading input files…")
df_a = pd.read_excel(FILE_A, dtype=str)
df_b = pd.read_excel(FILE_B, dtype=str)

# Normalise column names (strip whitespace)
df_a.columns = df_a.columns.str.strip()
df_b.columns = df_b.columns.str.strip()

print(f"  Linguist A columns : {list(df_a.columns)}")
print(f"  AI B columns       : {list(df_b.columns)}")
print(f"  Linguist A rows    : {len(df_a)}")
print(f"  AI B rows          : {len(df_b)}")

# Build column lookup (case-insensitive)
def col_lookup(df, name):
    """Return the actual column name in df that matches 'name' (case-insensitive)."""
    mapping = {c.lower(): c for c in df.columns}
    return mapping.get(name.lower())

# ── Compute per-language comparison data ─────────────────────────────────────
lang_results = {}   # lang -> list of row-dicts
lang_summary = []   # for Sheet 1

num_rows = min(len(df_a), len(df_b), 100)

for lang in LANGUAGES:
    col_a = col_lookup(df_a, lang)
    col_b = col_lookup(df_b, lang)
    col_en_a = col_lookup(df_a, "English")

    if col_a is None:
        print(f"  WARNING: '{lang}' column not found in Linguist A — skipping")
        continue
    if col_b is None:
        print(f"  WARNING: '{lang}' column not found in AI B — skipping")
        continue

    rows = []
    for i in range(num_rows):
        src  = str(df_a[col_en_a].iloc[i]).strip() if col_en_a else ""
        t_a  = str(df_a[col_a].iloc[i]).strip()
        t_b  = str(df_b[col_b].iloc[i]).strip()

        # Handle NaN / "nan"
        if t_a.lower() == "nan": t_a = ""
        if t_b.lower() == "nan": t_b = ""
        if src.lower() == "nan": src = ""

        exact  = t_a.lower() == t_b.lower()
        sim    = fuzz.token_sort_ratio(t_a, t_b)          # 0–100 float
        sim    = round(float(sim), 1)

        if exact:
            status = "Identical"
        elif sim >= 80:
            status = "Similar"
        else:
            status = "Divergent"

        rows.append({
            "Row #":            i + 1,
            "English":          src,
            "Linguist A":       t_a,
            "AI B":             t_b,
            "Exact Match":      "Yes" if exact else "No",
            "Similarity Score": sim,
            "Status":           status,
        })

    lang_results[lang] = rows

    # Summary stats
    scores       = [r["Similarity Score"] for r in rows]
    exact_count  = sum(1 for r in rows if r["Exact Match"] == "Yes")
    avg_sim      = round(sum(scores) / len(scores), 1) if scores else 0
    min_sim      = round(min(scores), 1) if scores else 0
    lt50         = sum(1 for s in scores if s < 50)
    b50_80       = sum(1 for s in scores if 50 <= s < 80)
    ge80         = sum(1 for s in scores if s >= 80)

    lang_summary.append({
        "Language":                 lang,
        "Exact Match Count":        exact_count,
        "Exact Match %":            round(exact_count / len(rows) * 100, 1),
        "Avg Similarity Score":     avg_sim,
        "Min Similarity Score":     min_sim,
        "Strings with score < 50":  lt50,
        "Strings with score 50–80": b50_80,
        "Strings with score ≥ 80":  ge80,
    })

    print(f"  {lang:12s}  exact={exact_count}/100  avg_sim={avg_sim}  <50:{lt50}  50-80:{b50_80}  >=80:{ge80}")

# ── Write output workbook ─────────────────────────────────────────────────────
print(f"\nWriting {OUTPUT} …")

with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    df_summary = pd.DataFrame(lang_summary)
    df_summary.to_excel(writer, sheet_name="Summary", index=False)

    ws_sum = writer.sheets["Summary"]
    # Header style
    for cell in ws_sum[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    # Data rows style
    for row_idx, row in enumerate(ws_sum.iter_rows(min_row=2), start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
        for cell in row:
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")

    # Auto-width
    for i in range(1, len(df_summary.columns) + 1):
        auto_width(ws_sum, i)

    ws_sum.freeze_panes = "A2"
    ws_sum.row_dimensions[1].height = 22

    # ── Sheets 2-6: per language ──────────────────────────────────────────────
    COLS = ["Row #", "English", "Linguist A", "AI B", "Exact Match",
            "Similarity Score", "Status"]
    SIM_COL  = COLS.index("Similarity Score") + 1   # 1-based
    STAT_COL = COLS.index("Status") + 1

    for lang in LANGUAGES:
        if lang not in lang_results:
            continue

        rows = lang_results[lang]
        df_lang = pd.DataFrame(rows, columns=COLS)
        df_lang.to_excel(writer, sheet_name=lang, index=False)

        ws = writer.sheets[lang]

        # Header
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER
        ws.row_dimensions[1].height = 22

        # Data rows — conditional formatting on Score & Status columns
        for row_idx, row_data in enumerate(rows, start=2):
            sim    = row_data["Similarity Score"]
            status = row_data["Status"]
            fill   = ALT_FILL if row_idx % 2 == 0 else PatternFill()
            cf     = score_fill(sim)

            for col_idx, cell in enumerate(ws[row_idx], start=1):
                cell.border = BORDER
                cell.alignment = Alignment(
                    horizontal="center" if col_idx not in (2, 3, 4) else "left",
                    vertical="center",
                    wrap_text=True if col_idx in (2, 3, 4) else False,
                )
                if col_idx in (SIM_COL, STAT_COL):
                    cell.fill = cf
                    cell.font = Font(bold=True)
                else:
                    cell.fill = fill

        # Auto-width
        for i, col_name in enumerate(COLS, start=1):
            if col_name in ("English", "Linguist A", "AI B"):
                ws.column_dimensions[get_column_letter(i)].width = 40
            else:
                auto_width(ws, i, min_w=12)

        ws.freeze_panes = "A2"

print("\nDone!")
print(f"Output saved to: {OUTPUT}")

# ── Stdout summary ────────────────────────────────────────────────────────────
print("\n" + "=" * 72)
print(f"{'TRANSLATION COMPARISON SUMMARY':^72}")
print("=" * 72)
header = (f"{'Language':<14} {'Exact':>6} {'Exact%':>7} {'AvgSim':>7} "
          f"{'MinSim':>7} {'<50':>5} {'50-80':>6} {'≥80':>5}")
print(header)
print("-" * 72)
for s in lang_summary:
    print(
        f"{s['Language']:<14} "
        f"{s['Exact Match Count']:>6} "
        f"{s['Exact Match %']:>6.1f}% "
        f"{s['Avg Similarity Score']:>7.1f} "
        f"{s['Min Similarity Score']:>7.1f} "
        f"{s['Strings with score < 50']:>5} "
        f"{s['Strings with score 50–80']:>6} "
        f"{s['Strings with score ≥ 80']:>5}"
    )
print("=" * 72)
print(f"\nTotal languages compared : {len(lang_summary)}")
print(f"Rows compared per language: {num_rows}")

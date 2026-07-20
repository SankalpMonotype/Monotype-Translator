#!/usr/bin/env python3
"""
import_glossary.py — Import linguist-approved ES glossary and translation memory
into the project knowledge base.

Usage:
    # Import EN-ES vocabulary glossary (appends to knowledge/glossary.md):
    python import_glossary.py --glossary "~/Downloads/NEW EN-ES Monotype Glossary Update.xlsx"

    # Import approved translation files into TM (appends to knowledge/tm.md):
    python import_glossary.py --tm-dir "~/Downloads/ES files for Agile POC/ES approved translated files/"

    # Both at once:
    python import_glossary.py
        --glossary "~/Downloads/NEW EN-ES Monotype Glossary Update.xlsx"
        --tm-dir "~/Downloads/ES files for Agile POC/ES approved translated files/"

Backs up each modified file as <file>.bak before writing.
Always commit the .bak alongside the change so the previous state is recoverable.
"""

import argparse
import glob
import os
import re
import shutil
import sys

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

KNOWLEDGE_DIR = os.path.join(os.path.dirname(__file__), "knowledge")
GLOSSARY_MD   = os.path.join(KNOWLEDGE_DIR, "glossary.md")
TM_MD         = os.path.join(KNOWLEDGE_DIR, "tm.md")

# These English strings should never be "translated" — they are brand / product names
# or technical identifiers that stay verbatim in all languages.
_UNTRANSLATABLE = {
    "monotype", "myfonts", "fonts.com", "monotype fonts", "monotype ai",
    "mosaic", "skyfonts", "anyword", "adobe fonts", "google fonts",
    "adobe creative cloud", "adobe creative suite", "arial", "helvetica",
    "api", "url", "html", "css", "id", "uuid", "pdf", "svg", "png", "jpg",
    "gcid", "ai", "saas",
}

# Strings that look like internal-only content (skip from TM)
_TM_SKIP_PATTERNS = [
    r"^dear\s",            # email salutations
    r"^estimado",          # Spanish salutations
    r"weekly pre-delivery", # internal reports
    r"automated update \(mondays",
    r"please find below",
    r"accounts (with|marked|in)",
]
_TM_SKIP_RE = re.compile("|".join(_TM_SKIP_PATTERNS), re.IGNORECASE)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _backup(path: str) -> str:
    bak = path + ".bak"
    shutil.copy2(path, bak)
    print(f"  Backed up {os.path.basename(path)} → {os.path.basename(bak)}")
    return bak


def _load_existing_text(*paths: str) -> str:
    text = ""
    for p in paths:
        if os.path.exists(p):
            with open(p, encoding="utf-8") as f:
                text += f.read().lower()
    return text


def _is_trivial(en: str, es: str) -> bool:
    """Return True for entries that add no value to the knowledge base."""
    en_lower = en.lower().strip()
    es_lower = es.lower().strip()
    # Skip entries where the translation equals the source (untranslatable)
    if en_lower == es_lower:
        return True
    # Skip brand/product names that are already handled
    if en_lower in _UNTRANSLATABLE:
        return True
    # Skip very short tokens (single characters, numbers)
    if len(en_lower) <= 1:
        return True
    return False


# ---------------------------------------------------------------------------
# Glossary import
# ---------------------------------------------------------------------------

def import_glossary(xlsx_path: str) -> None:
    print(f"\n=== Glossary import: {os.path.basename(xlsx_path)} ===")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        en = str(row[0]).strip() if row[0] else ""
        es = str(row[1]).strip() if row[1] else ""
        source = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        if en and es and en != "None" and es != "None":
            rows.append((en, es, source))
    wb.close()
    print(f"  Read {len(rows)} entries from Excel")

    # Load existing knowledge to find already-covered terms
    existing_text = _load_existing_text(
        GLOSSARY_MD,
        os.path.join(KNOWLEDGE_DIR, "es_ES_guide.md"),
    )

    # Filter
    new_entries = []
    skip_trivial = skip_existing = 0
    for en, es, source in rows:
        if _is_trivial(en, es):
            skip_trivial += 1
            continue
        if en.lower() in existing_text:
            skip_existing += 1
            continue
        new_entries.append((en, es, source))

    print(f"  Skipped trivial:    {skip_trivial}")
    print(f"  Skipped (existing): {skip_existing}")
    print(f"  New entries:        {len(new_entries)}")

    if not new_entries:
        print("  Nothing new to add — glossary.md is already up to date.")
        return

    # Group by source category for readability
    by_source: dict[str, list[tuple[str, str]]] = {}
    for en, es, source in new_entries:
        cat = source if source else "General"
        by_source.setdefault(cat, []).append((en, es))

    # Build the markdown block to append
    lines = [
        "",
        "",
        "---",
        "",
        "## Spanish (es-ES) Extended Vocabulary — Agile POC Glossary",
        "",
        "Linguist-approved Castilian Spanish vocabulary from the Monotype Agile POC glossary.",
        "Sourced from professional translators (RWS agency) and Monotype internal reviewers.",
        "Use these terms as the approved Castilian equivalent — do NOT use Latin American variants.",
        "",
    ]

    # Most useful categories first
    preferred_order = [
        "MONOTYPE", "AGILE", "INTERFAZ", "Glosarios tipográficos",
        "FUENTE", "FUNDICIÓN", "MONOTYPE TAGS", "TM", "RWS", "General",
    ]
    ordered_cats = sorted(
        by_source.keys(),
        key=lambda c: preferred_order.index(c) if c in preferred_order else len(preferred_order)
    )

    for cat in ordered_cats:
        entries = by_source[cat]
        if not entries:
            continue
        lines.append(f"### {cat}")
        lines.append("")
        lines.append("| English | Spanish (es-ES) |")
        lines.append("|---------|-----------------|")
        for en, es in sorted(entries, key=lambda x: x[0].lower()):
            # Escape pipe characters in cell content
            en_safe = en.replace("|", "\\|")
            es_safe = es.replace("|", "\\|")
            lines.append(f"| {en_safe} | {es_safe} |")
        lines.append("")

    block = "\n".join(lines)

    # Backup then append
    _backup(GLOSSARY_MD)
    with open(GLOSSARY_MD, "a", encoding="utf-8") as f:
        f.write(block)

    print(f"  Appended {len(new_entries)} entries to knowledge/glossary.md")


# ---------------------------------------------------------------------------
# TM import
# ---------------------------------------------------------------------------

def import_tm(tm_dir: str) -> None:
    print(f"\n=== TM import from: {tm_dir} ===")

    files = sorted(glob.glob(os.path.join(tm_dir, "*.xlsx")))
    print(f"  Found {len(files)} XLSX files")

    # Load existing TM to avoid duplicates
    existing_text = _load_existing_text(TM_MD)

    all_pairs: dict[str, str] = {}  # en -> es_ES
    for fpath in files:
        fname = os.path.basename(fpath)
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True)
            ws = wb.active
            headers = [
                str(c.value).strip().lower() if c.value else ""
                for c in next(ws.iter_rows(min_row=1, max_row=1))
            ]
            en_idx = next((i for i, h in enumerate(headers) if "english" in h), None)
            es_idx = next((i for i, h in enumerate(headers) if "spanish" in h or "español" in h), None)

            if en_idx is None or es_idx is None:
                print(f"  SKIP {fname}: missing English or Spanish column")
                wb.close()
                continue

            added = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                en = str(row[en_idx]).strip() if row[en_idx] else ""
                es = str(row[es_idx]).strip() if row[es_idx] else ""
                if not en or not es or en == "nan" or es == "nan":
                    continue
                if _is_trivial(en, es):
                    continue
                if _TM_SKIP_RE.search(en):
                    continue
                # Skip if already in TM
                if en.lower() in existing_text:
                    continue
                if en not in all_pairs:
                    all_pairs[en] = es
                    added += 1
            print(f"  {fname[:60]:<62} +{added}")
            wb.close()
        except Exception as exc:
            print(f"  ERROR {fname}: {exc}")

    print(f"\n  Total new EN→ES pairs to add: {len(all_pairs)}")
    if not all_pairs:
        print("  Nothing new to add — tm.md is already up to date.")
        return

    # Build markdown block — group into smaller tables of ≤50 rows each
    # so the TM stays readable. Add a clear section header.
    lines = [
        "",
        "",
        "---",
        "",
        "## Spanish (es-ES) Approved Translations — Agile POC",
        "",
        "Linguist-approved English → Castilian Spanish pairs from the Agile POC translation batches.",
        "**Use these exact translations for matching strings — do NOT rephrase or paraphrase.**",
        "",
        "| English | es-ES |",
        "|---------|-------|",
    ]

    for en, es in sorted(all_pairs.items(), key=lambda x: x[0].lower()):
        en_safe = en.replace("|", "\\|")
        es_safe = es.replace("|", "\\|")
        lines.append(f"| {en_safe} | {es_safe} |")

    lines.append("")
    block = "\n".join(lines)

    # Backup then append
    _backup(TM_MD)
    with open(TM_MD, "a", encoding="utf-8") as f:
        f.write(block)

    print(f"  Appended {len(all_pairs)} pairs to knowledge/tm.md")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--glossary", metavar="XLSX", help="EN-ES glossary Excel file to import")
    parser.add_argument("--tm-dir", metavar="DIR", help="Directory of approved translation XLSX files")
    args = parser.parse_args()

    if not args.glossary and not args.tm_dir:
        parser.print_help()
        sys.exit(1)

    if args.glossary:
        path = os.path.expanduser(args.glossary)
        if not os.path.exists(path):
            print(f"ERROR: {path} not found", file=sys.stderr)
            sys.exit(1)
        import_glossary(path)

    if args.tm_dir:
        path = os.path.expanduser(args.tm_dir)
        if not os.path.isdir(path):
            print(f"ERROR: {path} is not a directory", file=sys.stderr)
            sys.exit(1)
        import_tm(path)

    print("\nDone. Commit glossary.md, tm.md, and their .bak backups together.")


if __name__ == "__main__":
    main()

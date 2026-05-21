import hashlib
import json
import os
import re
from pathlib import Path
from crewai.tools import tool

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ---------------------------------------------------------------------------
# Column alias map — order-independent language column detection
# ---------------------------------------------------------------------------
COLUMN_ALIASES: dict[str, list[str]] = {
    "fr":     ["french", "fr", "français", "francais"],
    "de":     ["german", "de", "deutsch"],
    "pt_BR":  ["portuguese", "pt", "pt-br", "pt_br", "português", "portugues"],
    "ja":     ["japanese", "ja", "jp", "日本語"],
    "es_ES":  ["spanish", "es", "es-es", "es_es", "es-419", "es_ES", "español", "espanol"],
    "en":     ["english", "en", "source", "original", "string", "text"],
}

TARGET_LANGS = ["fr", "de", "pt_BR", "ja", "es_ES"]


def _normalise_translation(text: str, lang: str) -> str:
    """
    Apply deterministic post-processing to a translated string before writing to Excel.

    Rules applied (all languages):
      - Replace Unicode ellipsis U+2026 (…) with ASCII triple-dot (...).

    Rules applied (French only):
      - Replace non-breaking space U+00A0 immediately before ?, !, :, ;
        with a regular space (U+0020). The reference files use a regular space
        in these positions; the LLM often inserts the typographic non-breaking
        space, causing an otherwise-correct translation to fail exact matching.
    """
    # Universal: ellipsis normalisation
    text = text.replace('\u2026', '...')

    # French: non-breaking space before punctuation → regular space
    if lang == "fr":
        text = re.sub(r'\u00a0([?!:;])', r' \1', text)

    return text


def _detect_header_row(ws) -> int:
    """
    Scan rows 1-10 to find the first row where any cell matches a known language alias.
    Returns the 1-based row number, or 1 as fallback.
    """
    all_aliases = {alias for aliases in COLUMN_ALIASES.values() for alias in aliases}
    for row_idx in range(1, 11):
        row_values = [str(cell.value or "").strip().lower() for cell in ws[row_idx]]
        if any(v in all_aliases for v in row_values):
            return row_idx
    return 1


def _build_column_map(header_cells) -> dict[str, int]:
    """
    Returns a dict mapping canonical language keys (fr, de, pt_BR, ja, es_ES, en)
    to 0-based column indices.
    """
    col_map: dict[str, int] = {}
    for col_idx, cell in enumerate(header_cells):
        if cell.value is None:
            continue
        normalised = str(cell.value).strip().lower()
        for canonical, aliases in COLUMN_ALIASES.items():
            if normalised in aliases and canonical not in col_map:
                col_map[canonical] = col_idx
    return col_map


# ---------------------------------------------------------------------------
# Shared constants + internal write helper (used by Tool 0b and Tool 2)
# ---------------------------------------------------------------------------

REVIEWED_TRANSLATIONS_PATH = os.path.join("outputs", "reviewed_translations.json")

_LANG_HEADER_LABELS: dict[str, str] = {
    "fr": "French", "de": "German",
    "pt_BR": "Portuguese", "ja": "Japanese", "es_ES": "Spanish",
}
_META_KEYS = {"row_index", "english", "reviewer_note"}

# Extra normalisation for language keys the LLM may return with alternate spellings.
# Maps lower-cased aliases → canonical codes used in col_1based.
_LANG_KEY_ALIASES: dict[str, str] = {
    "pt":           "pt_BR",
    "pt-br":        "pt_BR",
    "pt_br":        "pt_BR",
    "es":           "es_ES",
    "es-419":       "es_ES",
    "es-es":        "es_ES",
    "french":       "fr",
    "german":       "de",
    "portuguese":   "pt_BR",
    "japanese":     "ja",
    "spanish":      "es_ES",
}


def _write_entries_to_excel(ws, entries: list[dict], header_row_idx: int) -> dict:
    """Write translation entries into an already-loaded worksheet.

    Adds missing language columns to the header row if needed, writes each
    entry's translations at the row_index specified, normalises text, and
    resets row heights so Excel auto-fits after the new content.

    Safety guards (skip + log rather than raise):
      - row_index ≤ header_row_idx  → never overwrite the header
      - row_index > ws.max_row      → never create phantom rows beyond the file
      - row has no English source   → never write to blank-source rows

    Returns a dict with rows_written, cells_written, skipped_rows, errors.
    """
    header_cells = list(ws.iter_rows(
        min_row=header_row_idx, max_row=header_row_idx, values_only=False
    ))[0]
    col_map = _build_column_map(header_cells)
    col_1based = {lang: idx + 1 for lang, idx in col_map.items()}

    en_col = col_1based.get("en")          # 1-based; None if no English column found
    max_data_row = ws.max_row              # upper bound — no phantom rows beyond this

    # Collect language keys, normalising LLM aliases (e.g. "pt" → "pt_BR")
    raw_langs = list(dict.fromkeys(
        k for entry in entries for k in entry if k not in _META_KEYS
    ))
    active_langs = list(dict.fromkeys(
        _LANG_KEY_ALIASES.get(k.lower(), k) for k in raw_langs
    ))

    # Map any normalised lang keys that are not yet in col_1based.
    # Try alias resolution first; only create a new column as a last resort.
    next_col = ws.max_column + 1
    for lang in active_langs:
        if lang not in col_1based:
            # See if a known alias already maps this to an existing column
            resolved = _LANG_KEY_ALIASES.get(lang.lower(), lang)
            if resolved in col_1based:
                col_1based[lang] = col_1based[resolved]
            else:
                ws.cell(row=header_row_idx, column=next_col,
                        value=_LANG_HEADER_LABELS.get(lang, lang))
                col_1based[lang] = next_col
                next_col += 1

    # Build a reverse map: raw entry key → normalised lang code (for text lookup)
    raw_to_norm = {
        k: _LANG_KEY_ALIASES.get(k.lower(), k)
        for k in raw_langs
    }

    rows_written = 0
    cells_written = 0
    skipped_rows: list[dict] = []
    errors: list[str] = []

    for entry in entries:
        row_index = entry.get("row_index")
        if not row_index or not isinstance(row_index, int):
            skipped_rows.append({"entry": str(entry)[:100], "reason": "missing row_index"})
            continue

        # Guard 1: never write to the header row (or above it)
        if row_index <= header_row_idx:
            skipped_rows.append({
                "entry": str(entry)[:100],
                "reason": f"row_index {row_index} is the header row or above — skipped",
            })
            continue

        # Guard 2: never create phantom rows beyond the worksheet's current extent
        if row_index > max_data_row:
            skipped_rows.append({
                "entry": str(entry)[:100],
                "reason": (
                    f"row_index {row_index} exceeds file max row {max_data_row} "
                    f"— likely an LLM row-count error; skipped"
                ),
            })
            continue

        # Guard 3: never write to a row whose English source cell is blank
        if en_col is not None:
            en_val = ws.cell(row=row_index, column=en_col).value
            if not en_val or str(en_val).strip() == "":
                skipped_rows.append({
                    "entry": str(entry)[:100],
                    "reason": f"row {row_index} has no English source text — skipped",
                })
                continue

        wrote_any = False
        for raw_key in raw_langs:
            norm_lang = raw_to_norm[raw_key]
            # Prefer the normalised key in the entry; fall back to the raw key
            text = entry.get(norm_lang) or entry.get(raw_key, "")
            if not text or str(text).strip() == "":
                continue
            col = col_1based.get(norm_lang)
            if col is None:
                errors.append(f"Row {row_index}: No column for language '{norm_lang}'")
                continue
            ws.cell(row=row_index, column=col,
                    value=_normalise_translation(str(text).strip(), norm_lang))
            cells_written += 1
            wrote_any = True

        if wrote_any:
            rows_written += 1

    for r in range(header_row_idx + 1, ws.max_row + 1):
        ws.row_dimensions[r].height = None

    return {
        "rows_written": rows_written,
        "cells_written": cells_written,
        "skipped_rows": skipped_rows,
        "errors": errors,
    }


def _save_translated_excel(wb, excel_path: str) -> str:
    """Save workbook to outputs/{stem}_translated{ext} and return the output path."""
    p = Path(excel_path)
    outputs_dir = Path(os.getcwd()) / "outputs"
    outputs_dir.mkdir(exist_ok=True)
    output_path = str(outputs_dir / f"{p.stem}_translated{p.suffix}")
    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# Tool 0: read_reviewed_translations
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# Tool 0b: write_reviewed_translations_to_excel
# ---------------------------------------------------------------------------

@tool
def write_reviewed_translations_to_excel(excel_path: str) -> str:
    """
    Reads the reviewed translations from outputs/reviewed_translations.json and
    writes them directly into the master Excel file. Saves output as a new file
    with '_translated' appended to the stem, preserving the original.

    This tool handles the full read-and-write in one step so that the large JSON
    payload never needs to pass through the LLM context as a tool argument.

    Args:
        excel_path: Path to the source .xlsx file.

    Returns:
        A JSON string with keys: success, output_path, rows_written,
        cells_written, skipped_rows, errors.
    """
    if not OPENPYXL_AVAILABLE:
        return json.dumps({"error": "openpyxl not installed. Run: pip install openpyxl"})

    json_path = REVIEWED_TRANSLATIONS_PATH
    if not os.path.isabs(json_path):
        json_path = os.path.join(os.getcwd(), json_path)
    if not os.path.exists(json_path):
        return json.dumps({"error": f"Reviewed translations file not found: {json_path}"})

    try:
        raw = open(json_path, encoding="utf-8").read()
    except Exception as exc:
        return json.dumps({"error": f"Failed to read {json_path}: {exc}"})

    cleaned = re.sub(r"^```(?:json)?\s*", "", raw.strip(), flags=re.IGNORECASE)
    cleaned = re.sub(r"\s*```$", "", cleaned)

    try:
        entries = json.loads(cleaned)
    except json.JSONDecodeError as exc:
        return json.dumps({"error": f"Invalid JSON in reviewed_translations.json: {exc}"})

    if not isinstance(entries, list):
        return json.dumps({"error": "reviewed_translations.json must contain a JSON array"})
    if not os.path.exists(excel_path):
        return json.dumps({"error": f"File not found: {excel_path}"})

    # For batch processing: if a _translated.xlsx already exists from a prior batch,
    # load it (not the original) so previously written translations are preserved.
    p = Path(excel_path)
    outputs_dir = Path(os.getcwd()) / "outputs"
    translated_path = outputs_dir / f"{p.stem}_translated{p.suffix}"
    load_path = str(translated_path) if translated_path.exists() else excel_path

    try:
        wb = openpyxl.load_workbook(load_path)
        ws = wb.active
        result = _write_entries_to_excel(ws, entries, _detect_header_row(ws))
        output_path = _save_translated_excel(wb, excel_path)
        return json.dumps({"success": True, "output_path": output_path, **result},
                          ensure_ascii=False, indent=2)
    except Exception as exc:
        return json.dumps({"error": f"Failed to write Excel: {exc}"})


@tool
def read_reviewed_translations() -> str:
    """
    Reads the reviewed translations JSON file written by the review_task.
    Returns the raw JSON string ready to pass to write_translations_to_excel.
    """
    path = REVIEWED_TRANSLATIONS_PATH
    if not os.path.isabs(path):
        path = os.path.join(os.getcwd(), path)
    if not os.path.exists(path):
        return json.dumps({"error": f"Reviewed translations file not found: {path}"})
    try:
        with open(path, "r", encoding="utf-8") as f:
            return f.read()
    except Exception as exc:
        return json.dumps({"error": f"Failed to read {path}: {exc}"})


# ---------------------------------------------------------------------------
# Tool 1: read_excel_for_translation
# ---------------------------------------------------------------------------

@tool
def read_excel_for_translation(
    excel_path: str,
    row_limit: int = 0,
    target_languages: str = "",
) -> str:
    """
    Reads the Monotype translation Excel file and returns rows that have English
    content but are missing one or more of the requested language translations.

    If a previously translated output file (_translated.xlsx) exists in the outputs/
    directory, it is used as the working file so that already-translated rows are
    automatically skipped — enabling safe batch-by-batch processing of large files.

    Args:
        excel_path: Absolute or relative path to the source .xlsx file.
        row_limit: Maximum number of untranslated rows to return (0 = no limit).
                   Set to 75 when processing large files to stay within LLM output limits.
        target_languages: Comma-separated list of language codes to check for completeness
                          (e.g. "fr,de,pt_BR,ja,es_ES" or "ja"). Only languages in this
                          list are used to decide whether a row still needs translation.
                          If empty, all supported languages are checked.

    Returns:
        A JSON string with keys:
        - excel_path
        - header_row (1-based row number of the header)
        - column_map (maps "en", "fr", "de", "pt_BR", "ja", "es_ES" to Excel column letters)
        - rows_to_translate (array of row objects needing translation)
        - total_rows_scanned
        - rows_needing_translation
        - rows_already_translated
    """
    if not OPENPYXL_AVAILABLE:
        return json.dumps({"error": "openpyxl not installed. Run: pip install openpyxl"})

    if not os.path.exists(excel_path):
        return json.dumps({"error": f"File not found: {excel_path}"})

    # Use the in-progress translated file if it exists, so already-translated rows
    # are treated as complete and skipped automatically (enables batch processing).
    p = Path(excel_path)
    outputs_dir = Path(os.getcwd()) / "outputs"
    translated_path = outputs_dir / f"{p.stem}_translated{p.suffix}"
    working_path = str(translated_path) if translated_path.exists() else excel_path

    try:
        wb = openpyxl.load_workbook(working_path, data_only=True, read_only=True)
        ws = wb.active

        # Detect header row
        header_row_idx = _detect_header_row(ws)
        header_cells = list(ws.iter_rows(
            min_row=header_row_idx, max_row=header_row_idx, values_only=False
        ))[0]
        col_map = _build_column_map(header_cells)

        # Build human-readable column letter map for reporting
        def idx_to_col_letter(idx: int) -> str:
            # openpyxl uses 1-based; idx here is 0-based
            from openpyxl.utils import get_column_letter
            return get_column_letter(idx + 1)

        col_letter_map = {lang: idx_to_col_letter(idx) for lang, idx in col_map.items()}

        if "en" not in col_map:
            return json.dumps({
                "error": (
                    "Could not detect an English source column. "
                    f"Headers found: {[c.value for c in header_cells]}"
                )
            })

        en_idx = col_map["en"]

        # Determine which languages to check for completeness.
        # Only languages in target_languages (if specified) count as "missing".
        # This prevents single-language jobs from looping forever because other
        # language columns remain empty throughout the run.
        requested = (
            [l.strip() for l in target_languages.split(",") if l.strip()]
            if target_languages
            else TARGET_LANGS
        )

        rows_to_translate = []
        total_scanned = 0

        # Use enumerate to track the actual 1-based Excel row number.
        # current_row = header_row_idx + offset (1 for the first data row, etc.)
        # This correctly handles blank rows that are skipped — they still consume a row number.
        for row_offset, row in enumerate(
            ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=1
        ):
            current_row_index = header_row_idx + row_offset  # actual 1-based Excel row

            # Skip blank rows
            if not any(row):
                continue

            en_val = row[en_idx] if en_idx < len(row) else None
            if not en_val or str(en_val).strip() == "":
                continue

            total_scanned += 1
            english = str(en_val).strip()

            # Only include requested languages in the payload — unrequested columns
            # are never sent to the LLM, avoiding unnecessary token consumption.
            lang_values: dict[str, str] = {}
            missing_langs: list[str] = []
            for lang in requested:
                if lang not in col_map:
                    missing_langs.append(lang)
                    lang_values[lang] = ""
                else:
                    cell_val = row[col_map[lang]] if col_map[lang] < len(row) else None
                    text = str(cell_val).strip() if cell_val else ""
                    lang_values[lang] = text
                    if not text:
                        missing_langs.append(lang)

            entry = {
                "row_index": current_row_index,
                "english": english,
                **lang_values,
                "missing_languages": missing_langs,
            }

            if missing_langs:
                rows_to_translate.append(entry)

        wb.close()

        # Apply batch cap — return only the first row_limit untranslated rows.
        if row_limit and row_limit > 0:
            rows_to_translate = rows_to_translate[:row_limit]

        result = {
            "excel_path": excel_path,
            "header_row": header_row_idx,
            "column_map": col_letter_map,
            "rows_to_translate": rows_to_translate,
            "total_rows_scanned": total_scanned,
            "rows_needing_translation": len(rows_to_translate),
            "rows_already_translated": total_scanned - len(rows_to_translate),
        }

        return json.dumps(result, ensure_ascii=False, indent=2, default=str)

    except Exception as exc:
        return json.dumps({"error": f"Failed to read Excel: {exc}"})


# ---------------------------------------------------------------------------
# Tool 2: write_translations_to_excel
# ---------------------------------------------------------------------------

@tool
def write_translations_to_excel(excel_path: str, translations_json: str) -> str:
    """
    Writes reviewed translations back into the master Excel file.
    Saves the output as a new file with '_translated' appended to the stem,
    preserving the original file.

    Args:
        excel_path: Path to the source .xlsx file.
        translations_json: A JSON array string. Each element must have:
            - "row_index": int (1-based Excel row number)
            - "fr", "de", "pt_BR", "ja", "es_ES": translated strings

    Returns:
        A JSON string with keys: success, output_path, rows_written,
        cells_written, skipped_rows, errors.
    """
    if not OPENPYXL_AVAILABLE:
        return json.dumps({"error": "openpyxl not installed. Run: pip install openpyxl"})

    # Strip markdown fences that LLMs sometimes add
    cleaned = re.sub(r"^```(?:json)?\s*", "", translations_json.strip(), flags=re.IGNORECASE)
    cleaned = re.sub(r"\s*```$", "", cleaned)

    try:
        entries = json.loads(cleaned)
    except json.JSONDecodeError as exc:
        return json.dumps({"error": f"Invalid JSON: {exc}", "received": translations_json[:500]})

    if not isinstance(entries, list):
        return json.dumps({"error": "translations_json must be a JSON array"})
    if not os.path.exists(excel_path):
        return json.dumps({"error": f"File not found: {excel_path}"})

    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        result = _write_entries_to_excel(ws, entries, _detect_header_row(ws))
        output_path = _save_translated_excel(wb, excel_path)
        return json.dumps({"success": True, "output_path": output_path, **result},
                          ensure_ascii=False, indent=2)
    except Exception as exc:
        return json.dumps({"error": f"Failed to write Excel: {exc}"})


# ---------------------------------------------------------------------------
# Tool 3: read_brand_guidelines
# ---------------------------------------------------------------------------

@tool
def read_brand_guidelines(knowledge_dir: str) -> str:
    """
    Reads all brand guideline and glossary files from the knowledge directory.
    Supports .md and .txt files. Returns their combined content.

    Args:
        knowledge_dir: Path to the directory containing brand files
                       (e.g. "knowledge" or an absolute path).

    Returns:
        Combined text content of all files, separated by filename headers.
    """
    if not os.path.isabs(knowledge_dir):
        knowledge_dir = os.path.join(os.getcwd(), knowledge_dir)

    kp = Path(knowledge_dir)
    if not kp.exists():
        return (
            f"Knowledge directory not found: {knowledge_dir}\n"
            "Please create a 'knowledge' directory with brand_guidelines.md and glossary.md."
        )

    # Read in priority order
    priority = ["brand_guidelines.md", "tone_of_voice.md", "glossary.md"]
    parts: list[str] = []
    read_names: set[str] = set()

    def _read(file_path: Path) -> None:
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                parts.append(f"=== {file_path.name} ===\n{f.read()}")
            read_names.add(file_path.name)
        except Exception as exc:
            parts.append(f"=== {file_path.name} === (read error: {exc})")

    for name in priority:
        fp = kp / name
        if fp.exists():
            _read(fp)

    for fp in sorted(kp.iterdir()):
        if fp.name in read_names:
            continue
        if fp.suffix.lower() in (".md", ".txt") and fp.is_file():
            _read(fp)

    if not parts:
        return (
            f"No guideline files found in {knowledge_dir}.\n"
            "Add .md or .txt files (brand_guidelines.md, glossary.md) to this directory."
        )

    return f"[{len(parts)} file(s) loaded from {knowledge_dir}]\n\n" + "\n\n".join(parts)


# ---------------------------------------------------------------------------
# Tool 4: read_brand_context_cache
# ---------------------------------------------------------------------------

BRAND_CONTEXT_CACHE_PATH = os.path.join("outputs", "brand_context_cache.md")
BRAND_CONTEXT_HASH_PATH  = os.path.join("outputs", "brand_context_cache.hash")
_KNOWLEDGE_DIR_FOR_HASH  = "knowledge"


def _hash_knowledge_files(knowledge_dir: str) -> str:
    """Returns an MD5 hex digest of all .md/.txt files in knowledge_dir, sorted by name."""
    kp = Path(knowledge_dir)
    if not kp.exists():
        return ""
    h = hashlib.md5()
    for fp in sorted(kp.iterdir()):
        if fp.suffix.lower() in (".md", ".txt") and fp.is_file():
            h.update(fp.name.encode())
            h.update(fp.read_bytes())
    return h.hexdigest()


@tool
def read_brand_context_cache(knowledge_dir: str) -> str:
    """
    Returns the cached brand context document if the knowledge files have not changed
    since it was last saved. Returns an empty string if the cache is missing or stale.

    Args:
        knowledge_dir: Path to the knowledge directory (e.g. "knowledge").

    Returns:
        Cached brand context string, or empty string on cache miss.
    """
    if not os.path.isabs(knowledge_dir):
        knowledge_dir = os.path.join(os.getcwd(), knowledge_dir)

    cache_path = BRAND_CONTEXT_CACHE_PATH
    hash_path  = BRAND_CONTEXT_HASH_PATH
    if not os.path.isabs(cache_path):
        cache_path = os.path.join(os.getcwd(), cache_path)
    if not os.path.isabs(hash_path):
        hash_path  = os.path.join(os.getcwd(), hash_path)

    if not os.path.exists(cache_path) or not os.path.exists(hash_path):
        return ""

    current_hash = _hash_knowledge_files(knowledge_dir)
    try:
        saved_hash = Path(hash_path).read_text(encoding="utf-8").strip()
    except Exception:
        return ""

    if current_hash != saved_hash:
        return ""

    try:
        return Path(cache_path).read_text(encoding="utf-8")
    except Exception:
        return ""


# ---------------------------------------------------------------------------
# Tool 5: save_brand_context_cache
# ---------------------------------------------------------------------------

@tool
def save_brand_context_cache(content: str, knowledge_dir: str) -> str:
    """
    Saves the brand context document to a cache file together with a hash of the
    current knowledge files. Future calls to read_brand_context_cache will return
    this content as long as the knowledge files remain unchanged.

    Args:
        content: The full brand context document text to cache.
        knowledge_dir: Path to the knowledge directory (e.g. "knowledge").

    Returns:
        A confirmation message.
    """
    if not os.path.isabs(knowledge_dir):
        knowledge_dir = os.path.join(os.getcwd(), knowledge_dir)

    cache_path = BRAND_CONTEXT_CACHE_PATH
    hash_path  = BRAND_CONTEXT_HASH_PATH
    if not os.path.isabs(cache_path):
        cache_path = os.path.join(os.getcwd(), cache_path)
    if not os.path.isabs(hash_path):
        hash_path  = os.path.join(os.getcwd(), hash_path)

    os.makedirs(os.path.dirname(cache_path), exist_ok=True)

    current_hash = _hash_knowledge_files(knowledge_dir)
    try:
        Path(cache_path).write_text(content, encoding="utf-8")
        Path(hash_path).write_text(current_hash, encoding="utf-8")
        return f"Brand context cached successfully ({len(content)} chars, hash={current_hash[:8]}...)."
    except Exception as exc:
        return f"Failed to save brand context cache: {exc}"
